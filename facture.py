import sys
from datetime import datetime
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from ui_utils import create_button , create_label, titre, create_table
import os

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1100, 600)

        self.titre = titre("GESTION DES CLIENTS",x = 10,y=15, parent=Form)
        # Logo
        self.logo = create_button("",400, 200, 300, 300,parent=Form)
        self.logo.setObjectName("logo")
        sary = QtGui.QIcon()
        sary.addFile("logo.png", QtCore.QSize())
        self.logo.setIcon(sary)
        Form.setWindowIcon(sary)
        self.logo.setIconSize(QtCore.QSize(300, 300))

        # Tableau des factures
        
        
        self.table_factures = create_table(20, 70, 1000, 200, parent=Form, 
                     column_count=4,
                    headers=["Numéro Facture", "Client", "Montant Total", "Actions"],
                    )
        self.table_factures.setColumnWidth(0, 200)
        self.table_factures.setColumnWidth(1, 300)
        self.table_factures.setColumnWidth(2, 200)
        self.table_factures.setColumnWidth(3, 100)

        # Tableau du détail
        
        self.table_detail = create_table(20, 300, 550, 250, parent=Form, 
                     column_count=4,
                    headers=["Désignation", "Quantité", "Prix Unitaire", "Montant"],
                    )
        self.table_detail.setColumnWidth(0, 200)
        # Chargement initial
        self.load_factures()
        # Boutons
        self.btn_actualiser = create_button("Actualiser",20, 560, 150, 30, parent=Form)
        self.btn_actualiser.clicked.connect(self.load_factures)
        self.btn_gerer_client = create_button("Gérer Client", 190, 560, 150, 30, parent=Form)
        self.btn_gerer_client.clicked.connect(self.ouvrir_gestion_client)
        # Bouton Exporter vers Excel
        self.btn_imprimer = create_button("Exporter vers Excel", 900, 560, 180, 30, parent=Form, object_name="btn_imprimer")
        self.btn_imprimer.clicked.connect(self.exporter_facture_selectionnee)

    # Ajoutez cette nouvelle méthode à la classe Ui_Form :
    def ouvrir_gestion_client(self):
        """Ouvre l'interface de gestion des clients"""
        import subprocess
        subprocess.Popen([sys.executable, "gestion_client.py"])


    def load_factures(self):
        """Charge la liste des factures groupées"""
        self.db_manager = FactureDatabaseManager("data_base.db")
        factures = self.db_manager.fetch_factures_grouped()
        self.db_manager.close()

        self.table_factures.setRowCount(len(factures))
        
        for row, (num_facture, client, total) in enumerate(factures):
            self.table_factures.setItem(row, 0, QtWidgets.QTableWidgetItem(num_facture))
            self.table_factures.setItem(row, 1, QtWidgets.QTableWidgetItem(client))
            self.table_factures.setItem(row, 2, QtWidgets.QTableWidgetItem(f"{total:,} Ar"))
            
            btn = QtWidgets.QPushButton("Voir détail")
            btn.clicked.connect(lambda _, n=num_facture: self.show_facture_detail(n))
            self.table_factures.setCellWidget(row, 3, btn)

    def show_facture_detail(self, num_facture):
        """Affiche le détail d'une facture"""
        self.db_manager = FactureDatabaseManager("data_base.db")
        details = self.db_manager.fetch_facture_details(num_facture)
        self.db_manager.close()

        self.table_detail.setRowCount(len(details))
        
        for row, (designation, qte, prix, montant) in enumerate(details):
            self.table_detail.setItem(row, 0, QtWidgets.QTableWidgetItem(designation))
            self.table_detail.setItem(row, 1, QtWidgets.QTableWidgetItem(str(qte)))
            self.table_detail.setItem(row, 2, QtWidgets.QTableWidgetItem(f"{prix:,} Ar"))
            self.table_detail.setItem(row, 3, QtWidgets.QTableWidgetItem(f"{montant:,} Ar"))

        # Ajouter une ligne de total
        total = sum(d[3] for d in details)
        self.table_detail.setRowCount(len(details) + 1)
        self.table_detail.setItem(len(details), 0, QtWidgets.QTableWidgetItem("TOTAL"))
        self.table_detail.setItem(len(details), 3, QtWidgets.QTableWidgetItem(f"{total:,} Ar"))

    def exporter_facture_selectionnee(self):
        """Exporte la facture sélectionnée vers Excel"""
        selected_row = self.table_factures.currentRow()
        if selected_row >= 0:
            num_facture = self.table_factures.item(selected_row, 0).text()
            self.exporter_vers_excel(num_facture)
        else:
            QMessageBox.warning(None, "Erreur", "Veuillez sélectionner une facture à exporter.")

    def exporter_vers_excel(self, num_facture):
        """Exporte une facture vers un fichier Excel et l'ouvre automatiquement"""
        from converture_en_lettre import nombre_en_lettres
        import subprocess
        import platform
        
        # Récupérer les données
        self.db_manager = FactureDatabaseManager("data_base.db")
        details = self.db_manager.fetch_facture_details(num_facture)
        info_facture = self.db_manager.fetch_facture_info(num_facture)
        self.db_manager.close()
        
        if not info_facture or not details:
            QMessageBox.warning(None, "Erreur", "Aucune donnée à exporter pour cette facture.")
            return
        
        num_facture, client, date_facture, total = info_facture
        
        # Créer le classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Facture {num_facture}"
        
        # Ajouter le logo si disponible
        try:
            if os.path.exists("logo2.png"):
                logo = Image("logo2.png")
                logo.height = 150
                logo.width = 150
                ws.add_image(logo, "C1")
        except Exception as e:
            print(f"Erreur lors du chargement du logo: {e}")

        # Styles
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal='center')
        border = Border(bottom=Side(style='thin'), top=Side(style='thin'), 
                      right=Side(style='thin'), left=Side(style='thin'))
        money_format = '#,##0.00" Ar"'
        
        # En-tête
        ws.merge_cells('A2:B2')
        ws['A2'] = "FACTURE"
        ws['A2'].font = Font('Franklin Gothic Demi Cond', bold=True, size=36)
        ws['A2'].alignment = center_aligned
        
        ws.merge_cells('A3:D3')
        ws['A3'] = f"Facture N°: {num_facture}"
        ws['A3'].font = Font('Times New Roman', bold=True)
        
        ws['A4'] = f"Date: {date_facture}"
        ws['A5'] = f"Client: {client}"
        ws['A6'] = "Adresse: "
        
        # En-têtes colonnes
        headers = ["Désignation", "Quantité", "Prix Unitaire", "Montant"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_aligned
        
        # Données
        for row, (designation, qte, prix, montant) in enumerate(details, 8):
            ws.cell(row=row, column=1, value=designation).border = border
            ws.cell(row=row, column=2, value=qte).border = border
            ws.cell(row=row, column=2).alignment = center_aligned
            ws.cell(row=row, column=3, value=prix).number_format = money_format
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=montant).number_format = money_format
            ws.cell(row=row, column=4).border = border
            
        # Total
        total_row = len(details) + 8
        ws.cell(row=total_row, column=3, value="TOTAL:").font = bold_font
        ws.cell(row=total_row, column=4, value=total).font = bold_font
        ws.cell(row=total_row, column=4).number_format = money_format
        ws.cell(row=total_row, column=3).border = border
        ws.cell(row=total_row, column=4).border = border
        
        # Montant en lettres
        try:
            montant_lettres = nombre_en_lettres(int(total))
        except Exception as e:
            print(f"Erreur conversion montant: {e}")
            montant_lettres = f"{total} Ariary"
            
        ws.merge_cells(f'A{total_row+1}:D{total_row+1}')
        ws.cell(row=total_row+1, column=1, 
               value=f"Arrêtée la présente facture à la somme de: {montant_lettres} Ariary")
        
        # Signature
        ws.cell(row=total_row+3, column=4, value="Le Responsable").font = Font(bold=True,  underline="single")
        
        
        # Ajuster les colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Créer le dossier facture s'il n'existe pas
        if not os.path.exists("facture"):
            os.makedirs("facture")
        
        # Sauvegarder
        filename = f"Facture_{num_facture}_{client}.xlsx"
        try:
            filepath = f"facture/{filename}"
            wb.save(filepath)
            
            # Ouvrir le fichier Excel automatiquement
            try:
                if platform.system() == 'Darwin':       # macOS
                    subprocess.call(('open', filepath))
                elif platform.system() == 'Windows':    # Windows
                    os.startfile(os.path.abspath(filepath))
                else:                                   # linux variants
                    subprocess.call(('xdg-open', filepath))
                    
                QMessageBox.information(None, "Succès", 
                                    f"Facture exportée et ouverte:\n{filename}")
            except Exception as open_error:
                QMessageBox.information(None, "Succès avec remarque", 
                                    f"Facture exportée mais impossible de l'ouvrir automatiquement:\n"
                                    f"{filename}\n\n"
                                    f"Erreur: {str(open_error)}")
                
        except Exception as e:
            QMessageBox.critical(None, "Erreur", f"Erreur lors de l'export:\n{str(e)}")
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    from facture_db_manager import FactureDatabaseManager

    try:
        with open("style_dark.qss", "r") as file:
            app.setStyleSheet(file.read())
    except FileNotFoundError:
        print("Le fichier style.qss n'a pas été trouvé.")

    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec())